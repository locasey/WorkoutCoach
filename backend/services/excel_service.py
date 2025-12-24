from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

class ExcelService:
    def __init__(self):
        pass
    
    def create_workout_plan_excel(self, workout_plan):
        """
        Create an Excel file from a workout plan dictionary.
        Returns a BytesIO object containing the Excel file.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Workout Plan"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title row
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"Workout Plan: {workout_plan.get('goal', 'Training Plan')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Info row
        ws.merge_cells('A2:F2')
        info_cell = ws['A2']
        duration = workout_plan.get('duration_weeks', 'N/A')
        created = workout_plan.get('created_at', datetime.now().isoformat())
        info_cell.value = f"Duration: {duration} weeks | Created: {created}"
        info_cell.font = Font(size=10, italic=True)
        info_cell.alignment = Alignment(horizontal='center')
        
        # Column headers
        headers = ['Week', 'Day', 'Type', 'Distance (km)', 'Duration (min)', 'Pace', 'Notes']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        workouts = workout_plan.get('workouts', [])
        day_names = {1: 'Monday', 2: 'Tuesday', 3: 'Wednesday', 4: 'Thursday', 
                    5: 'Friday', 6: 'Saturday', 7: 'Sunday'}
        
        for row_num, workout in enumerate(workouts, 4):
            ws.cell(row=row_num, column=1, value=workout.get('week', ''))
            ws.cell(row=row_num, column=2, value=day_names.get(workout.get('day', 1), ''))
            ws.cell(row=row_num, column=3, value=workout.get('type', '').replace('_', ' ').title())
            
            # Distance or Duration
            distance = workout.get('distance_km')
            duration = workout.get('duration_minutes')
            if distance:
                ws.cell(row=row_num, column=4, value=distance)
            else:
                ws.cell(row=row_num, column=4, value='')
            
            if duration:
                ws.cell(row=row_num, column=5, value=duration)
            else:
                ws.cell(row=row_num, column=5, value='')
            
            ws.cell(row=row_num, column=6, value=workout.get('pace', ''))
            ws.cell(row=row_num, column=7, value=workout.get('notes', ''))
            
            # Apply borders to all cells in the row
            for col_num in range(1, 8):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border
                if col_num == 7:  # Notes column
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    cell.alignment = Alignment(vertical='center')
        
        # Auto-adjust column widths
        column_widths = [8, 12, 18, 15, 15, 15, 50]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A4'
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file

